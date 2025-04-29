import openpyxl
print(openpyxl.__version__)

wb = openpyxl.load_workbook('./data/test.xlsx')
print(wb)

sheetnames=wb.sheetnames
print(sheetnames)

sheet = wb['基本信息']
print(sheet)
print(wb.active)

cell = sheet['A4'] #创建一个cell对象
print(cell.value)
print(cell.row)
print(cell.column)
print(cell.coordinate)

print(sheet.cell(row=1, column=2).value)

for cell_row in sheet['A2':'E6']:
    for cell in cell_row:
        print(cell.coordinate,cell.value)

print(list(sheet.columns)[0])


for cell in list(sheet.columns)[0]:
    print(cell.value)


print(sheet.max_row,sheet.max_column)